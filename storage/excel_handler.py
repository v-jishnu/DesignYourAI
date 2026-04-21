"""
Excel handler for MCQ knowledge base storage.
"""

from pathlib import Path
from typing import List
import logging
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook, load_workbook
except ImportError:
    pd = None
    Workbook = None
    load_workbook = None

from config.schemas import MCQ


class ExcelHandler:
    """Manage Excel operations for MCQ knowledge base."""

    def __init__(self, excel_path: Path):
        """
        Initialize Excel handler.

        Args:
            excel_path: Path to Excel file
        """
        if pd is None or Workbook is None:
            raise ImportError("pandas and openpyxl required. Install with: pip install pandas openpyxl")

        self.excel_path = Path(excel_path)
        self.logger = logging.getLogger(__name__)
        self._initialize_excel()

    def _initialize_excel(self):
        """Create Excel file with schema if doesn't exist."""
        if not self.excel_path.exists():
            self.logger.info(f"Creating new Excel file: {self.excel_path}")

            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "MCQs"

            # Headers
            headers = [
                'Question_ID', 'Question_Text', 'Option_A', 'Option_B',
                'Option_C', 'Option_D', 'Correct_Answer', 'Explanation',
                'Category', 'Topic', 'Difficulty', 'Source',
                'Date_Added', 'Used_Status', 'Company', 'Job_Roles',
                'Image_Path', 'Image_URL', 'Has_Image', 'Image_Format'
            ]
            ws.append(headers)

            # Save
            self.excel_path.parent.mkdir(exist_ok=True, parents=True)
            wb.save(self.excel_path)

    def append_mcqs(self, mcqs: List[MCQ]) -> int:
        """
        Append MCQs to Excel file.

        Args:
            mcqs: List of MCQ objects

        Returns:
            Number of MCQs stored
        """
        if not mcqs:
            return 0

        self.logger.info(f"Storing {len(mcqs)} MCQs to Excel")

        try:
            # Convert to DataFrame
            data = [mcq.to_dict() for mcq in mcqs]
            df_new = pd.DataFrame(data)

            # Load existing data (native dtypes — no dtype=str)
            try:
                df_existing = pd.read_excel(self.excel_path)
            except:
                df_existing = pd.DataFrame()

            # Align columns — add any missing columns to either side
            for col in df_new.columns:
                if col not in df_existing.columns:
                    df_existing[col] = None
            for col in df_existing.columns:
                if col not in df_new.columns:
                    df_new[col] = None

            # Match column order
            df_new = df_new[df_existing.columns]

            # Append
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)

            # Save — if batch fails, retry row-by-row to skip only the bad MCQ
            try:
                df_combined.to_excel(self.excel_path, index=False)
            except Exception as batch_err:
                self.logger.warning(f"Batch save failed ({batch_err}), retrying row-by-row...")
                saved_count = 0
                for i, mcq in enumerate(mcqs):
                    try:
                        row_df = pd.DataFrame([mcq.to_dict()])
                        for col in df_existing.columns:
                            if col not in row_df.columns:
                                row_df[col] = None
                        row_df = row_df[df_existing.columns]
                        combined = pd.concat([df_existing, row_df], ignore_index=True)
                        combined.to_excel(self.excel_path, index=False)
                        df_existing = combined
                        saved_count += 1
                    except Exception as row_err:
                        self.logger.warning(f"Skipping MCQ {i+1} (illegal chars): {str(row_err)[:80]}")
                        continue

                self.logger.info(f"Row-by-row recovery: saved {saved_count}/{len(mcqs)} MCQs")
                self._create_backup()
                return saved_count

            # Create backup
            self._create_backup()

            self.logger.info(f"Successfully stored {len(mcqs)} MCQs")
            return len(mcqs)

        except Exception as e:
            self.logger.error(f"Error storing MCQs: {e}")
            return 0

    def load_all_mcqs(self) -> List[MCQ]:
        """
        Load all MCQs from Excel.

        Returns:
            List of MCQ objects
        """
        if not self.excel_path.exists():
            return []

        try:
            df = pd.read_excel(self.excel_path)
            mcqs = []

            for _, row in df.iterrows():
                try:
                    mcq = MCQ.from_dict(row.to_dict())
                    mcqs.append(mcq)
                except Exception as e:
                    self.logger.warning(f"Error loading MCQ: {e}")
                    continue

            return mcqs

        except Exception as e:
            self.logger.error(f"Error loading MCQs: {e}")
            return []

    def get_mcq_count(self) -> int:
        """Get total number of MCQs."""
        if not self.excel_path.exists():
            return 0

        try:
            df = pd.read_excel(self.excel_path)
            return len(df)
        except:
            return 0

    def _create_backup(self):
        """Create timestamped backup."""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_path = self.excel_path.parent / f"backup_{timestamp}_{self.excel_path.name}"

            import shutil
            shutil.copy2(self.excel_path, backup_path)

            # Keep only last 5 backups
            backups = sorted(self.excel_path.parent.glob("backup_*"))
            if len(backups) > 5:
                for old_backup in backups[:-5]:
                    old_backup.unlink()

        except Exception as e:
            self.logger.warning(f"Error creating backup: {e}")
